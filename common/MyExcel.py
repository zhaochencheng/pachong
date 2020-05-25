#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@author: zhaochencheng
@contact: 907779487@qq.com
@file: MyExcel.py
@time: 2020/5/24 13:56
@IDE:PyCharm
'''
import openpyxl
from openpyxl import Workbook
import xlrd
import os

class MyExcel():
    # excel操作类
    def __init__(self, excel_name, sheetname):
        self.excel_name = excel_name
        # 判断excel文件是否存在
        if os.path.exists(excel_name):
            self.wb = openpyxl.load_workbook(excel_name)
            sheets = self.wb.sheetnames
            if sheetname in sheets:
                self.ws = self.wb[sheetname]
            else:
                self.ws = self.wb.create_sheet(sheetname)
        else:
            self.wb = Workbook()
            self.wb.save(excel_name)
            # 使用load_workbook可以追加sheet
            self.wb = openpyxl.load_workbook(excel_name)
            self.ws = self.wb.create_sheet(sheetname)
    def add_data(self,data):
        # 以list形式 新增数据
        self.ws.append(data)
    def add_row_column(self, row, column, data):
        # 指定行列 新增数据
        self.ws.cell(row=row, column=column, value=data)
    def save_excel(self):
        # 保存excel
        self.wb.save(self.excel_name)
        self.wb.close()
    def get_RowscolumnNum(self):
        # 返回sheetname 最大行数 和 最大列数
        rows = self.ws.max_row
        column = self.ws.max_column
        return rows, column
    def getCellValue(self, row, column):
        # 返回该row column 单元格的数据
        return self.ws.cell(row=row, column=column).value
    def delete_column(self, column, amount=1):
        # 执行此方法后，需调用save_excel
        return self.ws.delete_cols(idx=column,amount=amount)
    def delete_row(self, row, amount=1):
        # 执行此方法后，需调用save_excel
        return self.ws.delete_rows(idx=row, amount=amount)

    def copy_to_excel(self, dist_excel, distsheet_name):
        # 将当前execl文件中sheetname 拷贝到 另一个excel中
        # xlsx copy to xlsx
        if os.path.exists(dist_excel):
            dist_wb = openpyxl.load_workbook(dist_excel)
        else:
            dist_wb = Workbook()
            dist_wb.save(dist_excel)
        dist_ws = dist_wb.create_sheet(distsheet_name)

        source_max_row = self.ws.max_row
        source_max_column = self.ws.max_column
        print(source_max_row, source_max_column)
        for m in range(1, source_max_row+1):
            for n in range(1, source_max_column+1): #char(97) =A
                # n = chr(n) # ASCII字符
                # i = "%s%d" % (n,m) # 单元格编号
                # print(i)
                cell = self.getCellValue(row=m, column=n) #获取源execl文件 单元格数据
                # print("cell:",cell)
                dist_ws.cell(row=m, column=n).value = cell # 复制到目的execl文件 单元格
        dist_wb.save(dist_excel) # 目的excel文件数据保存
        dist_wb.close() #关闭目的excel文件

    def xls_copy_xlsx(self, source_xls, source_sheetname):

        xls = xlrd.open_workbook(source_xls)
        table = xls.sheet_by_name(source_sheetname)
        rows = table.nrows
        cols = table.ncols
        print(rows, cols)

        for i in range(0, rows):
            for j in range(0, cols):
                # print(table.cell_value(i,j))
                self.ws.cell(i+1, j+1).value = table.cell_value(i,j)
        self.save_excel()