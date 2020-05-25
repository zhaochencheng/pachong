#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
@author: zhaochencheng
@contact: 907779487@qq.com
@file: Data_ope.py
@time: 2020/5/24 13:57
@IDE:PyCharm
'''

import requests
import os
import sys
import threading
import queue as Queue
from lxml import html
etree = html.etree
import openpyxl
from openpyxl import Workbook
import xlrd
import time


class MyExcel():
    # excel操作类
    def __init__(self, excel_name, sheetname):
        self.excel_name = excel_name
        # 判断excel文件是否存在
        if os.path.exists(excel_name):
            self.wb = openpyxl.load_workbook(excel_name)
            sheets = self.wb.sheetnames
            if sheetname in sheets:
                self.ws = self.wb[sheetname]
            else:
                self.ws = self.wb.create_sheet(sheetname)
        else:
            self.wb = Workbook()
            self.wb.save(excel_name)
            # 使用load_workbook可以追加sheet
            self.wb = openpyxl.load_workbook(excel_name)
            self.ws = self.wb.create_sheet(sheetname)
    def add_data(self,data):
        # 以list形式 新增数据
        self.ws.append(data)
    def add_row_column(self, row, column, data):
        # 指定行列 新增数据
        self.ws.cell(row=row, column=column, value=data)
    def save_excel(self):
        # 保存excel
        self.wb.save(self.excel_name)
        self.wb.close()
    def get_RowscolumnNum(self):
        # 返回sheetname 最大行数 和 最大列数
        rows = self.ws.max_row
        column = self.ws.max_column
        return rows, column
    def getCellValue(self, row, column):
        # 返回该row column 单元格的数据
        return self.ws.cell(row=row, column=column).value
    def delete_column(self, column, amount=1):
        # 执行此方法后，需调用save_excel
        return self.ws.delete_cols(idx=column,amount=amount)
    def delete_row(self, row, amount=1):
        # 执行此方法后，需调用save_excel
        return self.ws.delete_rows(idx=row, amount=amount)

    def copy_to_excel(self, dist_excel, distsheet_name):
        # 将当前execl文件中sheetname 拷贝到 另一个excel中
        # xlsx copy to xlsx
        if os.path.exists(dist_excel):
            dist_wb = openpyxl.load_workbook(dist_excel)
        else:
            dist_wb = Workbook()
            dist_wb.save(dist_excel)
        dist_ws = dist_wb.create_sheet(distsheet_name)

        source_max_row = self.ws.max_row
        source_max_column = self.ws.max_column
        print(source_max_row, source_max_column)
        for m in range(1, source_max_row+1):
            for n in range(1, source_max_column+1): #char(97) =A
                # n = chr(n) # ASCII字符
                # i = "%s%d" % (n,m) # 单元格编号
                # print(i)
                cell = self.getCellValue(row=m, column=n) #获取源execl文件 单元格数据
                # print("cell:",cell)
                dist_ws.cell(row=m, column=n).value = cell # 复制到目的execl文件 单元格
        dist_wb.save(dist_excel) # 目的excel文件数据保存
        dist_wb.close() #关闭目的excel文件

    def xls_copy_xlsx(self, source_xls, source_sheetname):

        xls = xlrd.open_workbook(source_xls)
        table = xls.sheet_by_name(source_sheetname)
        rows = table.nrows
        cols = table.ncols
        print(rows, cols)

        for i in range(0, rows):
            for j in range(0, cols):
                # print(table.cell_value(i,j))
                self.ws.cell(i+1, j+1).value = table.cell_value(i,j)
        self.save_excel()






def check_rule(list, condition, rate_limit):
    #校验规则
    for each in list:
        if condition in each:
            print("each:",each)
            suit_check_rule_total = 0
            new = []
            for Growth_rate in each[-4:]:
                new.append(Growth_rate.strip())
            if new.count("-") < 2:
                for rate in new:
                    if rate == "-":
                        pass
                    else:
                        # print(rate_limit)
                        if float(rate.strip("%")) >= float(rate_limit):
                            suit_check_rule_total += 1
            if suit_check_rule_total >= 3:
                return True
            else:
                return False
def crawler(threadName,q, rate):
    # 从队列里获取code 股票代码
    result = q.get(timeout=2)
    try:
        print("开始请求同花顺页面")
        url_forecast = "http://basic.10jqka.com.cn/%s/worth.html" % result
        header ={"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.108 Safari/537.36"}
        r = requests.get(url=url_forecast, headers=header)
        # result = str(r.content, encoding="GBK")
        # 业绩预测分析
        content = etree.HTML(str(r.content, encoding="GBK"))
        title = content.xpath("/html/head/title/text()")[0].split(" ")[0]
        print(title)
        if os.path.exists(title+".xlsx"):
            os.remove(title+".xlsx")
        table = content.xpath("//table[@class='m_table m_hl ggintro ggintro_1 organData']")
        print("开始获取mongo数据-进行分析")
        if table:
            table_list = []
            for row in table:
                thead_th = row.xpath("./thead/tr/th/text()")
                tbody_th = row.xpath("./tbody/tr/th/text()")
                tbody_td = row.xpath("./tbody/tr/td")
            table_list.append(thead_th)
            td_list=[]
            for each in tbody_td:
                if len(each.xpath(".//span/text()")) == 0 :
                    td = each.xpath("./text()")[0].replace("\n","").replace("\t","")
                else:
                    td = each.xpath(".//span/text()")[0].replace("\n","").replace("\t","")
                td_list.append(td)
            data = []
            for i in range(0, len(td_list), 6):
                data.append(td_list[i:i+6])
            for j in range(0, len(tbody_th)):
                data[j].insert(0, tbody_th[j])
            for k in data:
                table_list.append(k)
            # print(table_list)
            myexcel = MyExcel(excel_name=title+".xlsx",sheetname=title+"业绩预测")
            # 匹配规则 营业收入增长率中 后四项存在三项大于10% 或 净利润增长率 后四项存在三项大于10% 输出
            if check_rule(list=table_list, condition="营业收入增长率", rate_limit=rate) and check_rule(list=table_list, condition="净利润增长率", rate_limit=rate):
                for each in table_list:
                    myexcel.add_data(each)
                myexcel.save_excel()

                #  # 获取财务分析表格
                url_debt = "http://basic.10jqka.com.cn/api/stock/export.php?export=debt&type=report&code=%s"% result
                header = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.108 Safari/537.36"}
                r2 = requests.get(url=url_debt, headers=header, timeout=5)
                with open(title+"财务分析表格.xlsx", "wb") as f:
                    f.write(r2.content)
                myexcel2 = MyExcel(excel_name=title+".xlsx",sheetname=title+"财务分析")
                # 将下载的xls 复制到 xlsc目标excel中
                myexcel2.xls_copy_xlsx(source_xls=title+"财务分析表格.xlsx",source_sheetname="Worksheet")
                max_row, max_column = myexcel2.get_RowscolumnNum()
                not_year_col =[]
                for excel2_column in range(2, max_column+1):
                    cell_data = myexcel2.getCellValue(row=2, column=excel2_column)
                    # print(cell_data, type(cell_data))
                    if "-12-" not in cell_data:
                        not_year_col.append(excel2_column)
                # print(not_year_col)
                for del_col in range(len(not_year_col)):
                    # print(del_col)
                    myexcel2.delete_column(column=not_year_col[del_col]-del_col)
                    myexcel2.save_excel()
                myexcel2.save_excel()
                if os.path.exists(title+"财务分析表格.xlsx"):
                    os.remove(title+"财务分析表格.xlsx")
            else:
                myexcel.save_excel()

    except Exception as e:
        print(q.qsize(),threadName,"Error: ",e)




class myThread(threading.Thread):
    def __init__(self, name,q, rate):
        threading.Thread.__init__(self)
        self.name = name
        self.q = q
        self.rate = rate
    def run(self):
        print("Starting " + self.name)
        while True:
            try:
                crawler(self.name,self.q, self.rate)
            except:
                break
        print("Exiting " + self.name)


# class myThread2(threading.Thread):
#     def __init__(self,name,q):
#         threading.Thread.__init__(self)
#         self.name = name
#         self.q = q
#     def run(self):
#         print("Starting " + self.name)
#         while True:
#             try:
#                 crawler(self.name,self.q)
#             except:
#                 break
#         print("Exiting " + self.name)

if __name__ == '__main__':
    start = time.time()

    # 请求同花顺获取 业务预测页面数据
    code_file = "股票信息.txt"
    # # 读取要爬取的股票数据
    start = time.time()
    with open(code_file, "r") as f:
        code_list = f.readlines()
    with open("rule.txt", "r") as f:
        rate = f.readlines()[0].strip()

    num = 0
    # 创建5个线程名
    threadList = ["Thread-%s"% i for i in range(20)]

    # 设置队列长度
    workQueue = Queue.Queue(3000)

    # 线程池
    threads = []
    #创建新线程
    for tName in threadList:
        thread = myThread(tName,workQueue, rate=rate)
        thread.start()
        threads.append(thread)
    #将url填充到队列
    for result in code_list:
        # print("code", code)
        workQueue.put(result.strip())


    #等待所有线程完成
    for t in threads:
        t.join()

    end = time.time()
    # print('Queue多线程爬虫总时间为：',get_time-start)
    print('Queue多线程爬虫总时间为：',end-start)
    #
